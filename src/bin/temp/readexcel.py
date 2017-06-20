from openpyxl import load_workbook
from openpyxl import Workbook
import os


class Excelhelper(object):
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def get_worksheet_object(self, sheet_name):
        if not os.path.exists(self.excel_file_path):
            print('Excel file {} does not exist'.format(self.excel_file_path))
            raise IOError('Excel file {} does not exist'.format(self.excel_file_path))
        wb = load_workbook(self.excel_file_path)
        try:
            print('Trying to access sheet {} from workbook'.format(sheet_name))
            ws = wb.get_sheet_by_name(sheet_name)
        except KeyError:
            raise KeyError('{} does not have {} sheet'.format(self.excel_file_path, sheet_name))
        else:
            return ws

    def get_sheet_data_as_array_of_dict(self, sheet_name):
        """ This function will return array of dict, dict will contain header name as key and row val as value"""
        ws = self.get_worksheet_object(sheet_name)
        headers = [row.value.strip() for row in ws.rows[0]]
        if not len(headers):
            raise Exception('No data in the sheet {}'.format(sheet_name))
        ret_list = []
        for row in ws.rows[1:]:
            datadict = {}
            for i in range(len(row)):
                if isinstance(row[i].value, str):
                    datadict[headers[i]] = row[i].value.strip()
                else:
                    datadict[headers[i]] = row[i].value
            ret_list.append(datadict)
        return ret_list

    def get_sheet_data_as_dict_of_dict(self, sheet_name, column_name):
        ws = self.get_worksheet_object(sheet_name)
        headers = [row.value.strip() for row in ws.rows[0]]
        if not len(headers):
            raise Exception('No data in the sheet {}'.format(sheet_name))
        if column_name not in headers:
            raise Exception('column name [{}] passed to a function does not exist in sheet [{}]'.format(column_name, sheet_name))
        datadict = {}
        for row in ws.rows[1:]:
            internal_dict = {}
            for i in range(len(row)):
                internal_dict[headers[i]] = row[i].value
            datadict[internal_dict[column_name]] = internal_dict
        return datadict

    def write_to_excel(self, dict_of_array_of_list):
        """Sample data : data = {'key':[['sql', 'xxx'], ['data', 'deepak']]}"""
        try:
            print('Writing result to excel file')
            wb = Workbook(write_only=True)
            for key, data in dict_of_array_of_list.items():
                ws = wb.create_sheet()
                ws.title = key
                for item in data:
                    ws.append(item)
            wb.save(self.excel_file_path)
            print('succesfully written data to file {}'.format(self.excel_file_path))
        except Exception as e:
            print('exception while writing report to excel : {}'.format(e.message))
